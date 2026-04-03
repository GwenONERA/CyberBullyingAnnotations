"""
Generate HTML visualizations for the three annotated datasets.
Highlights emotion expression mode spans with 4 colours:
  - Désignée   : rgb(40, 80, 160)   — Blue
  - Comportementale : rgb(0, 130, 100) — Teal
  - Suggérée   : rgb(180, 100, 20)  — Amber
  - Montrée    : rgb(120, 50, 140)  — Violet
Filters out texts with no annotated emotion (Emo == 0).
"""

import json
import html
import re
from pathlib import Path
from collections import Counter

import openpyxl

# ── colour map ──────────────────────────────────────────────────────
MODE_COLORS = {
    "Désignée":        "rgba(40, 80, 160, 0.30)",
    "Comportementale": "rgba(0, 130, 100, 0.30)",
    "Suggérée":        "rgba(180, 100, 20, 0.30)",
    "Montrée":         "rgba(120, 50, 140, 0.30)",
}
MODE_COLORS_SOLID = {
    "Désignée":        "rgb(40, 80, 160)",
    "Comportementale": "rgb(0, 130, 100)",
    "Suggérée":        "rgb(180, 100, 20)",
    "Montrée":         "rgb(120, 50, 140)",
}

MODES = list(MODE_COLORS.keys())

# ── column renaming (aggregate.py uses unaccented names) ────────────
COL_RENAME = {
    "Designee":    "Désignée",
    "Montree":     "Montrée",
    "Suggeree":    "Suggérée",
    "Colere":      "Colère",
    "Degout":      "Dégoût",
    "Culpabilite": "Culpabilité",
    "Fierte":      "Fierté",
}

def normalise_record(row: dict) -> dict:
    """Rename unaccented aggregate.py columns to accented names.
    Also promote _span_details → spans_json when spans_json is absent."""
    out = {}
    for k, v in row.items():
        out[COL_RENAME.get(k, k)] = v
    # _span_details → spans_json fallback
    if "spans_json" not in out and "_span_details" in out:
        out["spans_json"] = out["_span_details"]
    return out

# ── helpers ─────────────────────────────────────────────────────────

def read_xlsx(path: str) -> tuple[list[str], list[tuple]]:
    """Return (headers, rows) from first sheet."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = [tuple(c.value for c in row) for row in ws.iter_rows(min_row=2)]
    wb.close()
    return headers, rows


def rows_to_dicts(headers, rows):
    return [normalise_record(dict(zip(headers, r))) for r in rows]


def highlight_text_with_spans(text: str, spans_json_str: str) -> str:
    """Return HTML with highlighted spans. Handles overlapping spans."""
    if not spans_json_str or spans_json_str == "[]":
        return html.escape(text)

    try:
        spans = json.loads(spans_json_str)
    except (json.JSONDecodeError, TypeError):
        return html.escape(text)

    if not spans:
        return html.escape(text)

    # Build list of (start, end, mode) using re.search for each span
    intervals = []
    for sp in spans:
        span_text = sp.get("span_text", "")
        mode = sp.get("mode", "")
        if not span_text or mode not in MODE_COLORS:
            continue
        # Find all occurrences
        start = 0
        st_lower = text.lower()
        sp_lower = span_text.lower()
        while True:
            idx = st_lower.find(sp_lower, start)
            if idx == -1:
                break
            intervals.append((idx, idx + len(span_text), mode))
            break  # only first occurrence per span entry

    if not intervals:
        return html.escape(text)

    # Merge overlapping intervals: assign each character its modes
    char_modes = [set() for _ in range(len(text))]
    for start, end, mode in intervals:
        for i in range(start, min(end, len(text))):
            char_modes[i].add(mode)

    # Build runs of identical mode-sets
    result = []
    i = 0
    while i < len(text):
        modes = frozenset(char_modes[i])
        j = i + 1
        while j < len(text) and frozenset(char_modes[j]) == modes:
            j += 1
        chunk = html.escape(text[i:j])
        if modes:
            # Use the first mode's colour for background; add tooltip with all modes
            # If multiple modes, blend by stacking
            sorted_modes = sorted(modes)
            bg = MODE_COLORS[sorted_modes[0]]
            border_bottom = ""
            if len(sorted_modes) > 1:
                border_bottom = f"border-bottom:2px solid {MODE_COLORS_SOLID[sorted_modes[1]]};"
            title = " + ".join(sorted_modes)
            result.append(
                f'<span class="hl" style="background:{bg};{border_bottom}" title="{title}">{chunk}</span>'
            )
        else:
            result.append(chunk)
        i = j

    return "".join(result)


def mode_badges(row: dict) -> str:
    """For emotexttokids (no spans), return coloured badges for active modes."""
    badges = []
    for mode in MODES:
        if row.get(mode) and int(row[mode]) == 1:
            c = MODE_COLORS_SOLID[mode]
            badges.append(
                f'<span class="badge" style="background:{c};">{mode}</span>'
            )
    return " ".join(badges)


def compute_stats(records: list[dict], has_spans: bool) -> dict:
    """Return distribution statistics."""
    total = len(records)
    mode_counts = {m: 0 for m in MODES}
    emo_counts = Counter()
    emo_labels = [
        "Colère", "Dégoût", "Joie", "Peur", "Surprise", "Tristesse",
        "Admiration", "Culpabilité", "Embarras", "Fierté", "Jalousie", "Autre",
    ]
    for r in records:
        for m in MODES:
            if r.get(m) and int(r[m]) == 1:
                mode_counts[m] += 1
        for e in emo_labels:
            if r.get(e) and int(r[e]) == 1:
                emo_counts[e] += 1
    return {
        "total": total,
        "mode_counts": mode_counts,
        "emo_counts": dict(emo_counts),
    }


# ── HTML template ───────────────────────────────────────────────────

HTML_HEADER = """\
<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="utf-8"/>
<title>{title}</title>
<style>
  body {{
    font-family: "Segoe UI", Arial, sans-serif;
    font-size: 11px;
    line-height: 1.25;
    margin: 20px;
    color: #222;
  }}
  h1 {{ font-size: 18px; margin-bottom: 4px; }}
  h2 {{ font-size: 14px; margin: 12px 0 4px; }}
  .stats-grid {{
    display: flex; gap: 24px; flex-wrap: wrap;
    margin-bottom: 12px;
  }}
  .stat-box {{
    background: #f5f5f5; border-radius: 6px; padding: 8px 14px;
  }}
  .stat-box h3 {{ margin: 0 0 4px; font-size: 12px; }}
  .bar-row {{
    display: flex; align-items: center; gap: 4px;
    margin: 1px 0; font-size: 10px;
  }}
  .bar-label {{ width: 110px; text-align: right; }}
  .bar {{
    height: 12px; border-radius: 2px; min-width: 2px;
  }}
  .legend {{
    display: flex; gap: 14px; margin: 8px 0 10px;
    font-size: 11px; align-items: center;
  }}
  .legend-item {{
    display: inline-flex; align-items: center; gap: 3px;
  }}
  .legend-swatch {{
    display: inline-block; width: 14px; height: 14px;
    border-radius: 2px;
  }}
  .text-row {{
    padding: 3px 0;
    border-bottom: 1px solid #eee;
  }}
  .text-idx {{
    color: #999; font-size: 9px; margin-right: 6px;
    display: inline-block; width: 36px; text-align: right;
  }}
  .hl {{
    border-radius: 2px;
    padding: 0 1px;
  }}
  .badge {{
    display: inline-block;
    color: #fff;
    font-size: 9px;
    padding: 1px 5px;
    border-radius: 3px;
    margin-left: 3px;
  }}
  .meta {{
    color: #777; font-size: 9px; margin-left: 6px;
  }}
  #filter-bar {{
    margin: 8px 0;
    display: flex; gap: 8px; align-items: center; flex-wrap: wrap;
  }}
  #filter-bar label {{
    font-size: 11px; cursor: pointer;
  }}
  #filter-bar input[type=checkbox] {{ margin-right: 2px; }}
  #search-box {{
    font-size: 11px; padding: 2px 6px; width: 220px;
  }}
  .toggle-btn {{
    font-size: 10px; padding: 2px 8px; border: 1px solid #bbb;
    border-radius: 3px; cursor: pointer; background: #fff; color: #444;
  }}
  .toggle-btn.active {{
    background: #e0e0e0; border-color: #999;
  }}
  .char-name {{ font-size: 10px; font-weight: bold; }}
  .char-role {{ color: #777; font-size: 9px; margin-left: 6px; }}
  .hide-names .char-name {{ display: none; }}
  .hide-roles .char-role {{ display: none; }}
</style>
</head>
<body>
<h1>{title}</h1>
<p style="font-size:11px;color:#666;">{subtitle}</p>
"""

HTML_FOOTER = """\
<script>
function applyFilters() {
  var checks = document.querySelectorAll('.mode-filter');
  var activeModes = [];
  checks.forEach(function(cb) { if(cb.checked) activeModes.push(cb.value); });
  var search = (document.getElementById('search-box') || {}).value || '';
  search = search.toLowerCase();
  var rows = document.querySelectorAll('.text-row');
  var shown = 0;
  rows.forEach(function(row) {
    var modes = (row.dataset.modes || '').split(',');
    var modeOk = activeModes.length === 0 || activeModes.some(function(m){ return modes.indexOf(m) >= 0; });
    var textOk = !search || row.textContent.toLowerCase().indexOf(search) >= 0;
    if (modeOk && textOk) { row.style.display=''; shown++; }
    else { row.style.display='none'; }
  });
  document.getElementById('shown-count').textContent = shown;
}
document.querySelectorAll('.mode-filter').forEach(function(cb){ cb.addEventListener('change', applyFilters); });
var sb = document.getElementById('search-box');
if(sb) sb.addEventListener('input', applyFilters);
var btnName = document.getElementById('toggle-names');
var btnRole = document.getElementById('toggle-roles');
if(btnName) btnName.addEventListener('click', function() {
  document.body.classList.toggle('hide-names');
  this.classList.toggle('active');
});
if(btnRole) btnRole.addEventListener('click', function() {
  document.body.classList.toggle('hide-roles');
  this.classList.toggle('active');
});
</script>
</body>
</html>
"""


def bar_chart_html(counts: dict, color_map: dict | None = None, max_width: int = 200) -> str:
    """Horizontal bar chart as HTML."""
    if not counts:
        return ""
    max_val = max(counts.values()) if counts.values() else 1
    lines = []
    for label, val in sorted(counts.items(), key=lambda x: -x[1]):
        w = int(val / max_val * max_width) if max_val else 0
        c = (color_map or {}).get(label, "#888")
        lines.append(
            f'<div class="bar-row">'
            f'<span class="bar-label">{html.escape(label)}</span>'
            f'<div class="bar" style="width:{w}px;background:{c};"></div>'
            f'<span>{val}</span></div>'
        )
    return "\n".join(lines)


def legend_html() -> str:
    parts = []
    for mode, color in MODE_COLORS_SOLID.items():
        parts.append(
            f'<span class="legend-item">'
            f'<span class="legend-swatch" style="background:{color};"></span>'
            f'{mode}</span>'
        )
    return '<div class="legend">' + " ".join(parts) + "</div>"


def filter_bar_html(total: int, show_name_role: bool = False) -> str:
    checks = []
    for mode in MODES:
        c = MODE_COLORS_SOLID[mode]
        checks.append(
            f'<label style="color:{c};font-weight:600;">'
            f'<input type="checkbox" class="mode-filter" value="{mode}" checked> {mode}</label>'
        )
    name_role_btns = ""
    if show_name_role:
        name_role_btns = (
            '<button id="toggle-names" class="toggle-btn" title="Masquer/afficher les noms">Noms</button>'
            '<button id="toggle-roles" class="toggle-btn" title="Masquer/afficher les rôles">Rôles</button>'
        )
    return (
        f'<div id="filter-bar">'
        f'{"".join(checks)}'
        f'{name_role_btns}'
        f'<input type="text" id="search-box" placeholder="Rechercher dans les textes…">'
        f'<span style="font-size:10px;color:#999;">Affichés: <span id="shown-count">{total}</span> / {total}</span>'
        f'</div>'
    )


# ── generators ──────────────────────────────────────────────────────

def generate_emotexttokids_html(records, out_path):
    """emotexttokids has no spans_json → show badges for modes."""
    stats = compute_stats(records, has_spans=False)
    title = "EmoTextToKids — Visualisation des modes d'expression"
    subtitle = f"{stats['total']} textes avec au moins une émotion annotée (textes sans émotion filtrés)"

    parts = [HTML_HEADER.format(title=title, subtitle=subtitle)]
    parts.append('<div class="stats-grid">')
    parts.append('<div class="stat-box"><h3>Distribution des modes</h3>')
    parts.append(bar_chart_html(stats["mode_counts"], MODE_COLORS_SOLID))
    parts.append("</div>")
    parts.append('<div class="stat-box"><h3>Distribution des émotions</h3>')
    emo_colors = {e: "#888" for e in stats["emo_counts"]}
    parts.append(bar_chart_html(stats["emo_counts"], emo_colors))
    parts.append("</div></div>")

    parts.append(legend_html())
    parts.append(filter_bar_html(stats["total"]))
    parts.append("<h2>Textes</h2>")

    for i, r in enumerate(records):
        active_modes = [m for m in MODES if r.get(m) and int(r[m]) == 1]
        modes_str = ",".join(active_modes)
        text_esc = html.escape(str(r["TEXT"]))
        badges = mode_badges(r)
        emos = [e for e in ["Colère", "Dégoût", "Joie", "Peur", "Surprise",
                            "Tristesse", "Admiration", "Culpabilité", "Embarras",
                            "Fierté", "Jalousie", "Autre"]
                if r.get(e) and int(r[e]) == 1]
        emo_str = ", ".join(emos)
        parts.append(
            f'<div class="text-row" data-modes="{modes_str}">'
            f'<span class="text-idx">{i+1}</span>'
            f'{text_esc}{badges}'
            f'<span class="meta">[{emo_str}]</span>'
            f"</div>"
        )

    parts.append(HTML_FOOTER)
    Path(out_path).write_text("\n".join(parts), encoding="utf-8")
    print(f"  → {out_path}  ({stats['total']} texts)")


def generate_spans_html(records, out_path, dataset_name):
    """homophobie / obésité with spans_json → highlight spans in text."""
    stats = compute_stats(records, has_spans=True)
    title = f"{dataset_name} — Visualisation des modes d'expression"
    subtitle = f"{stats['total']} textes avec au moins une émotion annotée (textes sans émotion filtrés)"

    parts = [HTML_HEADER.format(title=title, subtitle=subtitle)]
    parts.append('<div class="stats-grid">')
    parts.append('<div class="stat-box"><h3>Distribution des modes</h3>')
    parts.append(bar_chart_html(stats["mode_counts"], MODE_COLORS_SOLID))
    parts.append("</div>")
    parts.append('<div class="stat-box"><h3>Distribution des émotions</h3>')
    emo_colors = {e: "#888" for e in stats["emo_counts"]}
    parts.append(bar_chart_html(stats["emo_counts"], emo_colors))
    parts.append("</div></div>")

    parts.append(legend_html())
    parts.append(filter_bar_html(stats["total"], show_name_role=True))
    parts.append("<h2>Textes</h2>")

        emos = [e for e in ["Colère", "Dégoût", "Joie", "Peur", "Surprise",
                            "Tristesse", "Admiration", "Culpabilité", "Embarras",
                            "Fierté", "Jalousie", "Autre"]
                if r.get(e) and int(r[e]) == 1]
        emo_str = ", ".join(emos)

        name = html.escape(str(r.get("NAME") or ""))
        role = html.escape(str(r.get("ROLE") or ""))

        parts.append(
            f'<div class="text-row" data-modes="{modes_str}">'
            f'<span class="text-idx">{i+1}</span>'
            f'<span class="char-name">{name}</span> '
            f'<span class="char-role">[{role}]</span> '
            f'{highlighted}'
            f'<span class="meta">[{emo_str}]</span>'
            f"</div>"
        )

    parts.append(HTML_FOOTER)
    Path(out_path).write_text("\n".join(parts), encoding="utf-8")
    print(f"  → {out_path}  ({stats['total']} texts)")


# ── main ────────────────────────────────────────────────────────────

def main():
    base = Path(r"C:\Users\gtsang\Annotation_")

    # 1. emotexttokids
    print("Processing emotexttokids…")
    h, rows = read_xlsx(base / "data" / "emotexttokids_gold_flat.xlsx")
    records = rows_to_dicts(h, rows)
    records = [r for r in records if r.get("Emo") and int(r["Emo"]) == 1]
    out = base / "outputs" / "viz_emotexttokids.html"
    generate_emotexttokids_html(records, out)

    # 2. homophobie
    print("Processing homophobie…")
    h, rows = read_xlsx(base / "outputs" / "homophobie" / "homophobie_annotations_gold_flat.xlsx")
    records = rows_to_dicts(h, rows)
    records = [r for r in records if r.get("Emo") and int(r["Emo"]) == 1]
    out = base / "outputs" / "homophobie" / "viz_homophobie.html"
    generate_spans_html(records, out, "Homophobie")

    # 3. obésité
    print("Processing obésité…")
    h, rows = read_xlsx(base / "outputs" / "obésité" / "obésité_annotations_gold_flat.xlsx")
    records = rows_to_dicts(h, rows)
    records = [r for r in records if r.get("Emo") and int(r["Emo"]) == 1]
    out = base / "outputs" / "obésité" / "viz_obésité.html"
    generate_spans_html(records, out, "Obésité")

    # 4. racisme (aggregate.py output — uses _span_details & unaccented cols)
    racisme_xlsx = base / "outputs" / "racisme" / "claude_racisme_aggregated.xlsx"
    if racisme_xlsx.exists():
        print("Processing racisme…")
        h, rows = read_xlsx(racisme_xlsx)
        records = rows_to_dicts(h, rows)
        records = [r for r in records if r.get("Emo") and int(r["Emo"]) == 1]
        out = base / "outputs" / "racisme" / "viz_racisme.html"
        generate_spans_html(records, out, "Racisme")
    else:
        print(f"⚠ Racisme XLSX introuvable : {racisme_xlsx}")
        print("  Lancez d'abord :  python scripts/annotation/aggregate.py ")
        print(f"    --input outputs/racisme/claude_racisme.jsonl ")
        print(f"    --output {racisme_xlsx}")

    print("Done.")


if __name__ == "__main__":
    main()
